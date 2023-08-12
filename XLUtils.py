import openpyxl

def getrowcount(file):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    return sheet.max_row

def getcolcount(file):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    return sheet.max_column

def readdata(file,rowno,colno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    return sheet.cell(rowno,colno).value

def writedata(file,rowno,colno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    sheet.cell(rowno,colno).value=data
    return workbook.save()

